from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Mot, Traduction, Langue
from .forms import MotForm, TraductionForm
from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from django.db.models import Count, F, Func, Value, CharField
from django.db.models.functions import Substr
from .models import Mot, Traduction, Langue
from collections import defaultdict
from django.db import IntegrityError
import os


def is_expert_or_admin(user):
    return user.is_authenticated and (user.user_type in [0, 1])


@login_required
@user_passes_test(is_expert_or_admin)

def importer_mots(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            mot, langue_nom, langue_code, audio_path, image_path = row[:5]  # Ajustez selon votre structure Excel

            try:
                # Vérifier si la langue existe, sinon la créer
                langue, created = Langue.objects.get_or_create(
                    code=langue_code,
                    defaults={'nom': langue_nom}
                )

                # Créer le mot
                nouveau_mot = Mot(mot=mot, langue=langue)

                # Gérer l'audio si présent
                if audio_path:
                    # Au lieu de vérifier si le fichier existe localement,
                    # vous pourriez enregistrer simplement le nom du fichier
                    nouveau_mot.prononciation_audio.name = audio_path

                # Gérer l'image si présente
                if image_path:
                    # Même approche que pour l'audio
                    nouveau_mot.image_illustrative.name = image_path

                nouveau_mot.save()

            except IntegrityError as e:
                messages.error(request, f"Erreur lors de l'importation du mot '{mot}': {str(e)}")
                continue

        messages.success(request, "Importation terminée !")
        return redirect('liste_mots')

    return render(request, 'dictionnaire/importer_mots.html')


@login_required
def dashboard_dictionnaire(request):
    # Statistiques de base
    total_mots = Mot.objects.count()
    total_traductions = Traduction.objects.count()
    total_langues = Langue.objects.count()

    # Récupérer les 5 derniers mots ajoutés
    derniers_mots = Mot.objects.order_by('-id')[:5]

    def simulate_monthly_data(queryset, total_count):
        data = list(queryset.values('id').order_by('-id'))
        monthly_data = defaultdict(int)
        if total_count > 0:
            chunk_size = max(1, total_count // 6)  # Assurez-vous que chunk_size est au moins 1
            for i, item in enumerate(data):
                month = min(5, i // chunk_size)  # Limite à 6 mois (0 à 5)
                monthly_data[month] += 1
        return [{'month': k, 'count': v} for k, v in monthly_data.items()]

    mots_par_mois = simulate_monthly_data(Mot.objects, total_mots)
    traductions_par_mois = simulate_monthly_data(Traduction.objects, total_traductions)

    # Données pour le graphique Traductions par Langue
    traductions_par_langue = Langue.objects.annotate(
        traductions_count=Count('mots__traductions_source', distinct=True) +
                          Count('mots__traductions_cible', distinct=True)
    ).values('nom', 'traductions_count').order_by('-traductions_count')[:5]

    # Données pour le graphique "Types" de Mots (basé sur la première lettre)
    types_de_mots = Mot.objects.annotate(
        first_letter=Substr('mot', 1, 1)
    ).values('first_letter').annotate(
        count=Count('id')
    ).order_by('-count')[:5]

    # Top 5 des mots les plus traduits
    mots_plus_traduits = Mot.objects.annotate(
        traductions_count=Count('traductions_source') + Count('traductions_cible')
    ).order_by('-traductions_count')[:5]

    context = {
        'total_mots': total_mots,
        'total_traductions': total_traductions,
        'total_langues': total_langues,
        'derniers_mots': derniers_mots,
        'mots_par_mois': mots_par_mois,
        'traductions_par_mois': traductions_par_mois,
        'traductions_par_langue': list(traductions_par_langue),
        'types_de_mots': list(types_de_mots),
        'mots_plus_traduits': mots_plus_traduits,
    }
    return render(request, 'dictionnaire/dashboard.html', context)

@login_required
def liste_mots(request):
    mots = Mot.objects.all().select_related('langue')
    total_mots = Mot.objects.count()
    total_traductions = Traduction.objects.count()
    total_langues = Langue.objects.count()
    derniers_mots = Mot.objects.order_by('-id')[:5].select_related('langue')
    langues = Langue.objects.all()

    context = {
        'mots': mots,
        'total_mots': total_mots,
        'total_traductions': total_traductions,
        'total_langues': total_langues,
        'derniers_mots': derniers_mots,
        'langues': langues,
    }
    return render(request, 'dictionnaire/liste_mots.html', context)
@login_required
def ajouter_mot(request):
    if request.method == 'POST':
        form = MotForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('liste_mots')
    else:
        form = MotForm()
    return render(request, 'dictionnaire/formulaire_mot.html', {'form': form})

@login_required
def modifier_mot(request, mot_id):
    mot = get_object_or_404(Mot, id=mot_id)
    if request.method == 'POST':
        form = MotForm(request.POST, request.FILES, instance=mot)
        if form.is_valid():
            form.save()
            return redirect('liste_mots')
    else:
        form = MotForm(instance=mot)
    return render(request, 'dictionnaire/formulaire_mot.html', {'form': form})


@login_required
def ajouter_traduction(request):
    if request.method == 'POST':
        form = TraductionForm(request.POST)
        if form.is_valid():
            traduction = form.save(commit=False)
            traduction.createur = request.user
            traduction.save()
            return redirect('liste_mots')
    else:
        form = TraductionForm()
    return render(request, 'dictionnaire/formulaire_traduction.html', {'form': form})

